from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from .models import Pedido, Cliente, Producto, DetallePedido


# aqui exportamos pedidos en pdf
@login_required
def exportar_pedidos_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedidos.pdf"'
    doc      = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles   = getSampleStyleSheet()
    pedidos  = Pedido.objects.all()
    elements.append(Paragraph("SISTEMA DE GESTIÓN", styles['Normal']))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("REPORTE DE PEDIDOS", styles['Title']))
    elements.append(Spacer(1, 10))

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements.append(Paragraph(f"Fecha de generación: {fecha}", styles['Normal']))
    elements.append(Paragraph(f"Total de pedidos: {pedidos.count()}", styles['Normal']))
    elements.append(Spacer(1, 20))
    data = [['ID', 'Cliente', 'Fecha', 'Estado']]
    for p in pedidos:
        data.append([
            p.id,
            p.cliente.nombre,
            str(p.fecha),
            p.estado
        ])

    table = Table(data, colWidths=[50, 150, 120, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Documento generado automáticamente por el sistema", styles['Italic']))

    doc.build(elements)
    return response


# aqui exportamos clientes en pdf
@login_required
def exportar_clientes_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clientes.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    clientes = Cliente.objects.all()
    elements.append(Paragraph("SISTEMA DE GESTIÓN", styles['Normal']))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("REPORTE DE CLIENTES", styles['Title']))
    elements.append(Spacer(1, 10))

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements.append(Paragraph(f"Fecha de generación: {fecha}", styles['Normal']))
    elements.append(Paragraph(f"Total de clientes: {clientes.count()}", styles['Normal']))
    elements.append(Spacer(1, 20))
    data = [['ID', 'Nombre', 'Correo', 'Teléfono']]
    for c in clientes:
        data.append([
            c.id,
            c.nombre,
            c.correo,
            c.telefono
        ])

    table = Table(data, colWidths=[50, 150, 200, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Documento generado automáticamente por el sistema", styles['Italic']))

    doc.build(elements)
    return response


# aqui exportamos productos en pdf
@login_required
def exportar_productos_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    productos = Producto.objects.all()
    elements.append(Paragraph("SISTEMA DE GESTIÓN", styles['Normal']))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("REPORTE DE PRODUCTOS", styles['Title']))
    elements.append(Spacer(1, 10))

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements.append(Paragraph(f"Fecha de generación: {fecha}", styles['Normal']))
    elements.append(Paragraph(f"Total de productos: {productos.count()}", styles['Normal']))
    elements.append(Spacer(1, 20))
    data = [['ID', 'Nombre', 'Precio', 'Stock']]
    for p in productos:
        data.append([
            p.id,
            p.nombre,
            str(p.precio),
            p.stock
        ])

    table = Table(data, colWidths=[50, 180, 100, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Documento generado automáticamente por el sistema", styles['Italic']))

    doc.build(elements)
    return response


# aqui exportamos detalles en pdf
@login_required
def exportar_detalles_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="detalles.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    detalles = DetallePedido.objects.all()
    elements.append(Paragraph("SISTEMA DE GESTIÓN", styles['Normal']))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("REPORTE DE TODOS LOS DETALLES", styles['Title']))
    elements.append(Spacer(1, 10))

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements.append(Paragraph(f"Fecha de generación: {fecha}", styles['Normal']))
    elements.append(Spacer(1, 20))
    data = [['Ped#', 'Cliente', 'Producto', 'Cant', 'Subtotal']]
    for d in detalles:
        data.append([
            d.pedido.id,
            d.pedido.cliente.nombre,
            d.producto.nombre,
            d.cantidad,
            str(d.subtotal)
        ])

    table = Table(data, colWidths=[40, 120, 180, 50, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Documento generado automáticamente por el sistema", styles['Italic']))

    doc.build(elements)
    return response


# aqui armamos un excel con varias hojas
@login_required
def exportar_todo_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reportes.xlsx'

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Pedidos"
    ws1.append(['ID', 'Cliente', 'Fecha', 'Estado'])
    for p in Pedido.objects.all():
        ws1.append([p.id, p.cliente.nombre, str(p.fecha), p.estado])
    ws2 = wb.create_sheet(title="Clientes")
    ws2.append(['ID', 'Nombre', 'Correo', 'Teléfono'])
    for c in Cliente.objects.all():
        ws2.append([c.id, c.nombre, c.correo, c.telefono])
    ws3 = wb.create_sheet(title="Productos")
    ws3.append(['ID', 'Nombre', 'Precio', 'Stock'])
    for p in Producto.objects.all():
        ws3.append([p.id, p.nombre, p.precio, p.stock])

    wb.save(response)
    return response


# aqui exportamos todos los detalles en excel
@login_required
def exportar_detalles_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=detalles.xlsx'

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Todos los Detalles"
    ws1.append(['ID Detalle', 'Pedido', 'Cliente', 'Producto', 'Cantidad', 'Subtotal'])
    for d in DetallePedido.objects.all():
        ws1.append([d.id, d.pedido.id, d.pedido.cliente.nombre, d.producto.nombre, d.cantidad, d.subtotal])

    wb.save(response)
    return response


# aqui sacamos pedidos con detalle en json
@login_required
def exportar_pedidos_detallado_json(request):
    pedidos = Pedido.objects.all()
    data = []
    for p in pedidos:
        detalles = []
        for d in p.detallepedido_set.all():
            detalles.append({
                'producto': d.producto.nombre,
                'cantidad': d.cantidad,
                'subtotal': float(d.subtotal)
            })
        data.append({
            'id': p.id,
            'cliente': p.cliente.nombre,
            'fecha': str(p.fecha),
            'estado': p.estado,
            'detalles': detalles
        })
    return JsonResponse(data, safe=False)


# aqui sacamos clientes en json
@login_required
def exportar_clientes_json(request):
    clientes = Cliente.objects.all()
    data = []
    for c in clientes:
        data.append({
            'id': c.id,
            'nombre': c.nombre,
            'correo': c.correo,
            'telefono': c.telefono
        })
    return JsonResponse(data, safe=False)


# aqui sacamos productos en json
@login_required
def exportar_productos_json(request):
    productos = Producto.objects.all()
    data = []
    for p in productos:
        data.append({
            'id': p.id,
            'nombre': p.nombre,
            'precio': float(p.precio),
            'stock': p.stock
        })
    return JsonResponse(data, safe=False)

